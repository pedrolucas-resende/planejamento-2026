"""
Script para extrair gráficos da aba OVERVIEW do Excel e salvar em data/images
Dependências: openpyxl, pillow
"""

import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import logging

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def export_charts_from_excel(
    excel_path: str = 'data/xlsx/overview_sales.xlsx',
    sheet_name: str = 'OVERVIEW',
    output_dir: str = 'data/images'
) -> None:
    """
    Extrai gráficos da aba OVERVIEW do Excel e salva em data/images
    
    Args:
        excel_path: Caminho para o arquivo Excel
        sheet_name: Nome da aba que contém os gráficos
        output_dir: Diretório de saída para as imagens
    """
    
    # Criar diretório de saída se não existir
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    # Verificar se arquivo existe
    if not os.path.exists(excel_path):
        logger.error(f"Arquivo não encontrado: {excel_path}")
        return
    
    try:
        # Carregar workbook
        wb = load_workbook(excel_path)
        
        # Verificar se aba existe
        if sheet_name not in wb.sheetnames:
            logger.error(f"Aba '{sheet_name}' não encontrada. Abas disponíveis: {wb.sheetnames}")
            return
        
        ws = wb[sheet_name]
        logger.info(f"Processando aba: {sheet_name}")
        
        # Extrair gráficos
        chart_count = 0
        if hasattr(ws, '_charts'):
            for idx, chart in enumerate(ws._charts, 1):
                # Gerar nome do arquivo baseado no título do gráfico ou índice
                chart_name = chart.title if chart.title else f"grafico_{idx}"
                # Sanitizar nome do arquivo
                chart_name = "".join(c if c.isalnum() or c in ('-', '_') else '_' for c in chart_name)
                
                output_path = os.path.join(output_dir, f"{chart_name}.png")
                
                logger.info(f"Gráfico encontrado: {chart_name}")
                # Nota: openpyxl não exporta gráficos diretamente
                # Você precisará exportar manualmente do Excel como PNG
                logger.warning("Para exportar gráficos do Excel, use: Clique direito no gráfico → 'Copiar' → Colar em editor de imagem")
                
                chart_count += 1
        
        if chart_count == 0:
            logger.info("Nenhum gráfico encontrado na aba.")
            logger.info("Dica: Exporte manualmente do Excel (Clique direito → 'Copiar' → Salve como imagem)")
        else:
            logger.info(f"Total de gráficos encontrados: {chart_count}")
        
        wb.close()
        
    except Exception as e:
        logger.error(f"Erro ao processar Excel: {e}")

def list_charts_in_overview(
    excel_path: str = 'data/xlsx/overview_sales.xlsx',
    sheet_name: str = 'OVERVIEW'
) -> list:
    """
    Lista os gráficos disponíveis na aba OVERVIEW
    
    Args:
        excel_path: Caminho para o arquivo Excel
        sheet_name: Nome da aba
        
    Returns:
        Lista com nomes dos gráficos
    """
    
    if not os.path.exists(excel_path):
        logger.error(f"Arquivo não encontrado: {excel_path}")
        return []
    
    try:
        wb = load_workbook(excel_path)
        
        if sheet_name not in wb.sheetnames:
            logger.error(f"Aba '{sheet_name}' não encontrada.")
            return []
        
        ws = wb[sheet_name]
        charts = []
        
        if hasattr(ws, '_charts'):
            for idx, chart in enumerate(ws._charts, 1):
                chart_name = chart.title if chart.title else f"Gráfico {idx}"
                charts.append(chart_name)
                logger.info(f"  {idx}. {chart_name}")
        
        wb.close()
        return charts
        
    except Exception as e:
        logger.error(f"Erro ao listar gráficos: {e}")
        return []

if __name__ == '__main__':
    import sys
    
    print("\n=== Exportador de Gráficos do Excel ===\n")
    
    # Listar gráficos disponíveis
    print("Gráficos disponíveis na aba 'OVERVIEW':")
    charts = list_charts_in_overview()
    
    if charts:
        print(f"\nTotal: {len(charts)} gráfico(s)")
        print("\nPara exportar os gráficos, siga estes passos:")
        print("1. Abra o arquivo: data/xlsx/overview_sales.xlsx")
        print("2. Vá para a aba: OVERVIEW")
        print("3. Para cada gráfico:")
        print("   - Clique direito no gráfico")
        print("   - Selecione 'Copiar'")
        print("   - Abra Paint, Preview ou outro editor de imagem")
        print("   - Cole (Ctrl+V ou Cmd+V)")
        print("   - Salve em: data/images/ com um nome descritivo")
        print("\nExemplos de nomes:")
        print("  - demanda_filial_temporal.png")
        print("  - demanda_produto_0km.png")
        print("  - demanda_capital_interior.png")
    else:
        print("Nenhum gráfico encontrado. Verifique se a aba 'OVERVIEW' existe e contém gráficos.")
    
    print("\n" + "="*50 + "\n")
